from openpyxl import load_workbook

import configparser
import time
import ftplib
import os
import mysql.connector
import sys

def getftpfiles(ftpcfg, remote_dir, local_dir):
	ftp = ftplib.FTP(ftpcfg['hostname'], ftpcfg['username'], ftpcfg['password'])
	ftp.cwd(remote_dir)

	#Get pending files
	filenames = ftp.nlst()
	for filename in filenames:
		host_file = os.path.join(local_dir, filename)
		try:
			with open(host_file, 'wb') as local_file:
				ftp.retrbinary('RETR ' + filename, local_file.write)
		except ftplib.error_perm:
			pass

	ftp.quit()
	return filenames

def getsheetdata(targetfile):
	wb = load_workbook(targetfile, read_only=True)
	sheets = set(wb.get_sheet_names()).intersection(target_sheets) #Get only present sheets
	status = {'status': False, 'sheets': {}}
	status_data = {}
	sheet_status = []
	filename = os.path.basename(targetfile)
	period = filename.split('-')[1].replace('.xlsx','').split('_')
	period_month = period[0]
	period_year = period[1]

	for sheet in sheets:
		#Get sheet data
		ws = wb[sheet]
		highest_row = ws.max_row
		highest_column = ws.max_column
		#Filter based on sheetname
		if sheet == 'Ordering Points':
			status_data[sheet] = getfacilityorderingsites(cfg, ws, highest_row, highest_column, filename, sheet)
		elif sheet == 'Current patients by ART site':
			status_data[sheet] = getfacilitypatients(cfg, ws, highest_row, highest_column, period_month, period_year, filename, sheet)
		elif sheet == 'Facility Cons by ARV Medicine':
			status_data[sheet] = getfacilityconsumption(cfg, ws, highest_row, highest_column, period_month, period_year, filename, sheet)
		elif sheet == 'Facility SOH by ARV Medicine':
			status_data[sheet] = getfacilitysoh(cfg, ws, highest_row, highest_column, period_month, period_year, filename, sheet)
		elif sheet == 'Stock Status':
			status_data[sheet] = getnationalmos(cfg, ws, highest_row, highest_column, period_year, filename, sheet)
		#Update sheet status
		sheet_status.append(status_data[sheet]['status'])
	#Set response
	if False not in sheet_status:
		status['status'] = True
	status['sheets'] = status_data
	wb.close()
	return status

def getfacilityorderingsites(cfg, ws, highest_row, highest_column, filename, sheet):
	status = {'status': True, 'message': 'Sheet Import Success!'}	
	maincfg = cfg['facility_ordering_sites']
	start_row = int(maincfg['first_row'])
	facility_col = str(maincfg['facility_col'])
	code_col = str(maincfg['code_col'])
	county_col = str(maincfg['county_col'])
	proc_name = str(maincfg['proc_name'])

	for i in range(start_row, highest_row):
		row = str(i)
		facility_name = str(ws[facility_col+row].value).replace("'", "").lower().lstrip().rstrip()
		facility_code = str(ws[code_col+row].value).replace("'", "").upper().lstrip().rstrip()
		county_name =str(ws[county_col+row].value).replace("'", "").lower().lstrip().rstrip()
		if(facility_name != 'None'):
			response = runproc(cfg, filename, sheet, proc_name, [facility_code, facility_name, county_name])
			if(response['status'] == False):
				status = {'status': False, 'message': 'Sheet Import Failed!'}

	return status

def runproc(cfg, filename, sheet, proc_name, proc_args):
	#Get database connection
	cnx = getdbconnection(cfg)
	cursor = cnx.cursor()
	try:
		result_args = cursor.callproc(proc_name, proc_args)
		#Commit changes and close connection
		cnx.commit()
		cursor.close()
		cnx.close()
		return {'status': True, 'message': result_args[1]}
	except Exception, e:
		logfile = cfg['main']['logs_dir']+'Failed_'+filename+'.log'
		message = "SHEET: "+ sheet + " ERROR: "+ str(e)+" PROCEDURE: CALL "+proc_name+"("+','.join("'{0}'".format(x) for x in proc_args)+");"
		writelog(logfile, message)
		return {'status': False, 'message': e}

def getdbconnection(cfg):
	#Database configuration
	dbcfg = {
		'user': cfg["database"]["username"],
		'password': cfg["database"]["password"],
		'host': cfg["database"]["hostname"],
		'database': cfg["database"]["dbname"],
		'port': cfg["database"]["port"]
	}

	return mysql.connector.connect(**dbcfg)

def writelog(logfile,logmsg):
	import logging

	logger = logging.getLogger(__name__)
	logger.setLevel(logging.INFO)
	#create a logfile handler
	handler = logging.FileHandler(logfile)
	handler.setLevel(logging.INFO)
	#create a logging format
	formatter = logging.Formatter('%(asctime)s %(message)s')
	handler.setFormatter(formatter)
	#add the handlers to the logger
	logger.addHandler(handler)
	#write message to logfile
	logger.info(logmsg)
	#remove handler
	logger.removeHandler(handler)

def getfacilitypatients(cfg, ws, highest_row, highest_column, period_month, period_year, filename, sheet):
	status = {'status': True, 'message': 'Sheet Import Success!'}	
	maincfg = cfg['facility_patients']
	start_row = int(maincfg['first_row'])
	mflcode_col = str(maincfg['mflcode_col'])
	id_col = str(maincfg['id_col'])
	first_col_index = int(maincfg['first_col_index'])
	regimen_row = int(maincfg['regimen_row'])
	proc_name = str(maincfg['proc_name'])
	cols = getexcelcolumns(first_col_index, highest_column)
	proc_sql = ''

	for i in range(start_row, highest_row):
		row = str(i)
		id_val = ws[id_col+row].value
		facility_code = str(ws[mflcode_col+row].value).replace("'", "").lstrip().rstrip()
		if(id_val):
			for col in cols:
				regimen_code = str(ws[col+str(regimen_row)].value).replace("'", "").lstrip().rstrip()
				if(regimen_code != 'None'):
					patient_total = str(ws[col+row].value)
					response = runproc(cfg, filename, sheet, proc_name, [facility_code, regimen_code, patient_total, period_month, period_year])
					if(response['status'] == False):
						status = {'status': False, 'message': 'Sheet Import Failed!'}

	return status

def getexcelcolumns(start_index, end_index):
	cols = []
	for col_index in range(start_index, end_index):
		col = getcolumnletter(col_index)
		cols.append(col)
	return cols

def getcolumnletter(num):
	letters = ''
	while num:
		mod = (num - 1) % 26
		letters += chr(mod + 65)
		num = (num - 1) // 26
	return ''.join(reversed(letters))

def getfacilityconsumption(cfg, ws, highest_row, highest_column, period_month, period_year, filename, sheet):
	status = {'status': True, 'message': 'Sheet Import Success!'}	
	maincfg = cfg['facility_consumption']
	start_row = int(maincfg['first_row'])
	id_col = str(maincfg['id_col'])
	mflcode_col = str(maincfg['mflcode_col'])
	first_col_index = int(maincfg['first_col_index'])
	cols = getexcelcolumns(first_col_index, highest_column)
	drug_row = str(maincfg['drug_row'])
	packsize_row = str(maincfg['packsize_row'])
	proc_name = str(maincfg['proc_name'])
	proc_sql = ''

	for i in range(start_row, highest_row):
		row = str(i)
		id_val = ws[id_col+row].value
		facility_code = str(ws[mflcode_col+row].value).replace("'", "").lstrip().rstrip()
		if(id_val):
			for col in cols:
				drug_name = str(ws[col+drug_row].value).replace("'", "").replace("  ", " ").lstrip().rstrip()
				packsize = str(ws[col+str(packsize_row)].value).replace("'", "").replace("  ", " ").lstrip().rstrip().lstrip().rstrip()
				if(drug_name and packsize and drug_name != 'None' and packsize != 'None'):
					consumption_total = str(ws[col+row].value)
					response = runproc(cfg, filename, sheet, proc_name, [facility_code, drug_name, packsize, period_year, period_month, consumption_total])
					if(response['status'] == False):
						status = {'status': False, 'message': 'Sheet Import Failed!'}

	return status

def getfacilitysoh(cfg, ws, highest_row, highest_column, period_month, period_year, filename, sheet):
	status = {'status': True, 'message': 'Sheet Import Success!'}	
	maincfg = cfg['facility_soh']
	start_row = int(maincfg['first_row'])
	id_col = str(maincfg['id_col'])
	mflcode_col = str(maincfg['mflcode_col'])
	first_col_index = int(maincfg['first_col_index'])
	cols = getexcelcolumns(first_col_index, highest_column)
	drug_row = str(maincfg['drug_row'])
	packsize_row = str(maincfg['packsize_row'])
	proc_name = str(maincfg['proc_name'])
	proc_sql = ''

	for i in range(start_row, highest_row):
		row = str(i)
		id_val = ws[id_col+row].value
		facility_code = str(ws[mflcode_col+row].value).lstrip().rstrip()
		if(id_val):
			for col in cols:
				drug_name = str(ws[col+str(drug_row)].value).replace("'", "").replace("  ", " ").lstrip().rstrip()
				packsize = str(ws[col+str(packsize_row)].value).replace("'", "").replace("  ", " ").lstrip().rstrip()
				if(drug_name and packsize and drug_name != 'None' and packsize != 'None'):
					soh_total = str(ws[col+row].value)
					response = runproc(cfg, filename, sheet, proc_name, [facility_code, drug_name, packsize, period_year, period_month, soh_total])
					if(response['status'] == False):
						status = {'status': False, 'message': 'Sheet Import Failed!'}

	return status

def getnationalmos(cfg, ws, highest_row, highest_column, period_year, filename, sheet):
	status = {'status': True, 'message': 'Sheet Import Success!'}	
	maincfg = cfg['national_mos']
	start_row = int(maincfg['first_row'])
	id_col = maincfg['id_col']
	drug_col = maincfg['drug_col']
	packsize_col = maincfg['packsize_col']
	first_col_index = int(maincfg['first_col_index'])
	cols = getexcelcolumns(first_col_index, highest_column+1)
	month_row = str(maincfg['month_row'])
	proc_name = str(maincfg['proc_name'])
	proc_sql = ''

	for i in range(start_row, highest_row):
		row = str(i)
		id_val = ws[id_col+row].value
		drug_name = str(ws[drug_col+row].value).replace("'", "").replace("  ", " ").lstrip().rstrip()
		packsize = str(ws[packsize_col+row].value).replace("'", "").replace("  ", " ").lstrip().rstrip()
		if(id_val):
			for index, col in enumerate(cols):
				period_month = str(ws[col+month_row].value).title().lstrip().rstrip()
				if(period_month != 'None'):
					#Set total indices
					issue_col = str(cols[index+int(maincfg['issue_index'])])
					soh_col = str(cols[index+int(maincfg['soh_index'])])
					supplier_col = str(cols[index+int(maincfg['supplier_index'])])
					received_col = str(cols[index+int(maincfg['received_index'])])
					#Get totals
					issues_total = str(ws[issue_col+row].value)
					soh_total = str(ws[soh_col+row].value)
					supplier_total = str(ws[supplier_col+row].value)
					received_total = str(ws[received_col+row].value)
					response = runproc(cfg, filename, sheet, proc_name, [drug_name, packsize, period_year, period_month, issues_total, soh_total, supplier_total, received_total])
					if(response['status'] == False):
						status = {'status': False, 'message': 'Sheet Import Failed!'}

	return status

def moveftpfile(ftpcfg, source, destination):
	ftp = ftplib.FTP(ftpcfg['hostname'], ftpcfg['username'], ftpcfg['password'])
	return ftp.rename(source, destination)

def uploadftpfile(ftpcfg, sourcedir, destinationdir, filename):
	ftp = ftplib.FTP(ftpcfg['hostname'], ftpcfg['username'], ftpcfg['password'])
	#Set params
	source = os.path.join(sourcedir, filename)
	destination = os.path.join(destinationdir, filename)
	#Upload file
	file = open(source,'r')
	ftp.storbinary('STOR '+destination, file)
	file.close()
	ftp.quit() 

if __name__ == '__main__':
	#Encode strings to utf-8
	reload(sys)
	sys.setdefaultencoding('utf-8')

	#Get configuration
	cfg = configparser.ConfigParser()
	cfg.read('config/properties.ini')

	#Get params 
	files_dir = cfg['main']['files_dir']
	logs_dir = cfg['main']['logs_dir']
	ftpusername = cfg['ftp']['username']
	ftppassword = cfg['ftp']['password']
	pending_dir = cfg['ftp']['pending_dir']
	completed_dir = cfg['ftp']['completed_dir']
	failed_dir = cfg['ftp']['failed_dir']
	target_sheets = cfg['main']['target_sheets'].split(',')
	response = {}

	#FTP configuration
	ftpcfg = {
		'hostname': cfg['ftp']['hostname'],
		'username': cfg['ftp']['username'],
		'password': cfg['ftp']['password']
	}

	#Get pending files
	for filename in getftpfiles(ftpcfg, pending_dir, files_dir):
		start = time.time()
		localfile = os.path.join(files_dir, filename)
		response[filename] = getsheetdata(localfile)
		#Remove localfile
		os.remove(localfile)
		#Move remotefile to /dir
		source = pending_dir+filename
		if(response[filename]['status']):
			destination = completed_dir+filename
			moveftpfile(ftpcfg, source, destination)
			print 'SUCCESS:',filename,'was processed in',time.strftime('%H:%M:%S', time.gmtime(time.time()-start))
		else:
			destination = failed_dir+filename
			moveftpfile(ftpcfg, source, destination)
			#Upload logfile to ftp
			logfile = 'Failed_'+filename+'.log'
			uploadftpfile(ftpcfg, logs_dir, failed_dir, logfile)
			#Remove local logfile
			os.remove(os.path.join(logs_dir, logfile))
			print 'ERROR:',filename,' failed to be processed in',time.strftime('%H:%M:%S', time.gmtime(time.time()-start))	